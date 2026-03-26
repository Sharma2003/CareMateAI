from fastapi import APIRouter, HTTPException, UploadFile, File, Query
from pydantic import BaseModel
from typing import Optional, List
from uuid import UUID
from database.core import DbSession
from auth.service import CurrentUser
from entities.Drug import Drug
from sqlalchemy import or_, func
import io

router = APIRouter(prefix="/drugs", tags=["drugs"])


class DrugCreate(BaseModel):
    name: str
    generic_name: Optional[str] = None
    brand_name: Optional[str] = None
    category: Optional[str] = None
    drug_class: Optional[str] = None
    common_dosages: Optional[str] = None
    route: Optional[str] = "oral"
    description: Optional[str] = None

    model_config = {"extra": "ignore"}


class DrugResponse(BaseModel):
    id: UUID
    name: str
    generic_name: Optional[str] = None
    brand_name: Optional[str] = None
    category: Optional[str] = None
    drug_class: Optional[str] = None
    common_dosages: Optional[str] = None
    route: Optional[str] = None
    description: Optional[str] = None
    usage_count: int = 0
    source: str = "manual"

    model_config = {"from_attributes": True}


@router.get("/search", response_model=List[DrugResponse])
def search_drugs(
    q: str = Query(..., min_length=1, description="Drug name to search"),
    limit: int = Query(20, le=50),
    db: DbSession = None,
    current_user: CurrentUser = None,
):
    """Search drugs from centralized DB by name or generic name."""
    term = f"%{q.lower()}%"
    results = (
        db.query(Drug)
        .filter(
            Drug.is_active == True,
            or_(
                func.lower(Drug.name).like(term),
                func.lower(Drug.generic_name).like(term),
                func.lower(Drug.brand_name).like(term),
            )
        )
        .order_by(Drug.usage_count.desc(), Drug.name)
        .limit(limit)
        .all()
    )
    return results


@router.get("/popular", response_model=List[DrugResponse])
def get_popular_drugs(
    limit: int = Query(50, le=200),
    db: DbSession = None,
    current_user: CurrentUser = None,
):
    """Get most popular/commonly used drugs."""
    results = (
        db.query(Drug)
        .filter(Drug.is_active == True)
        .order_by(Drug.usage_count.desc(), Drug.name)
        .limit(limit)
        .all()
    )
    return results


@router.post("", response_model=DrugResponse)
@router.post("/", response_model=DrugResponse, include_in_schema=False)
def add_drug(payload: DrugCreate, db: DbSession = None, current_user: CurrentUser = None):
    """Doctor manually adds a new drug to the database."""
    # Check for duplicate
    existing = db.query(Drug).filter(func.lower(Drug.name) == payload.name.lower()).first()
    if existing:
        return existing  # Return existing if already in DB
    drug = Drug(
        name=payload.name.strip(),
        generic_name=payload.generic_name,
        brand_name=payload.brand_name,
        category=payload.category,
        drug_class=payload.drug_class,
        common_dosages=payload.common_dosages,
        route=payload.route or "oral",
        description=payload.description,
        source="manual",
    )
    db.add(drug)
    db.commit()
    db.refresh(drug)
    return drug


@router.post("/increment-usage/{drug_name}")
def increment_usage(drug_name: str, db: DbSession = None, current_user: CurrentUser = None):
    """Track when a drug is prescribed (increments usage count)."""
    drug = db.query(Drug).filter(func.lower(Drug.name) == drug_name.lower()).first()
    if drug:
        drug.usage_count = (drug.usage_count or 0) + 1
        db.commit()
    return {"ok": True}


@router.post("/bulk-import")
async def bulk_import_drugs(
    file: UploadFile = File(...),
    db: DbSession = None,
    current_user: CurrentUser = None,
):
    """Import drugs from Excel (.xlsx) or CSV file."""
    filename = file.filename or ""
    content = await file.read()

    added = 0
    skipped = 0
    errors = []

    try:
        if filename.endswith(".csv"):
            import csv
            reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig")))
            rows = list(reader)
        elif filename.endswith(".xlsx") or filename.endswith(".xls"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(content))
                ws = wb.active
                headers = [str(cell.value or "").strip().lower() for cell in ws[1]]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append({headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)})
            except ImportError:
                raise HTTPException(status_code=400, detail="openpyxl not installed. Please use CSV format.")
        else:
            raise HTTPException(status_code=400, detail="Only .xlsx or .csv files supported")

        for i, row in enumerate(rows):
            # Accept common column name variations
            name = (row.get("name") or row.get("drug_name") or row.get("medicine") or "").strip()
            if not name:
                skipped += 1
                continue
            existing = db.query(Drug).filter(func.lower(Drug.name) == name.lower()).first()
            if existing:
                skipped += 1
                continue
            try:
                drug = Drug(
                    name=name,
                    generic_name=(row.get("generic_name") or row.get("generic") or "").strip() or None,
                    brand_name=(row.get("brand_name") or row.get("brand") or "").strip() or None,
                    category=(row.get("category") or "").strip() or None,
                    drug_class=(row.get("drug_class") or row.get("class") or "").strip() or None,
                    common_dosages=(row.get("common_dosages") or row.get("dosage") or row.get("dosages") or "").strip() or None,
                    route=(row.get("route") or "oral").strip(),
                    description=(row.get("description") or row.get("desc") or "").strip() or None,
                    source="import",
                )
                db.add(drug)
                added += 1
            except Exception as e:
                errors.append(f"Row {i+2}: {str(e)}")

        db.commit()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")

    return {
        "message": f"Import complete: {added} added, {skipped} skipped",
        "added": added,
        "skipped": skipped,
        "errors": errors[:10],  # first 10 errors
    }


@router.get("/download-template")
def download_template():
    """Download Excel template for bulk import."""
    from fastapi.responses import StreamingResponse
    csv_content = "name,generic_name,brand_name,category,drug_class,common_dosages,route,description\n"
    csv_content += "Paracetamol,Acetaminophen,Calpol,Analgesic,Non-opioid analgesic,\"500mg, 1000mg\",oral,Pain reliever and fever reducer\n"
    csv_content += "Amoxicillin,Amoxicillin,Amoxil,Antibiotic,Penicillin,\"250mg, 500mg\",oral,Broad-spectrum antibiotic\n"
    return StreamingResponse(
        iter([csv_content]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=drug_import_template.csv"}
    )


@router.post("/seed-database")
def seed_drug_database(db: DbSession = None, current_user: CurrentUser = None):
    """Seed the drug database with comprehensive Indian/International drug list. Call once after setup."""
    from drugs.seed_data import get_seed_drugs
    drugs_data = get_seed_drugs()
    added = 0
    skipped = 0
    for row in drugs_data:
        name, generic, brand, category, dosages, route, usage_count = row
        existing = db.query(Drug).filter(func.lower(Drug.name) == name.lower()).first()
        if existing:
            skipped += 1
            continue
        drug = Drug(
            name=name, generic_name=generic, brand_name=brand,
            category=category, common_dosages=dosages, route=route,
            usage_count=usage_count, source="seed",
        )
        db.add(drug)
        added += 1
    db.commit()
    return {"message": f"Seed complete: {added} added, {skipped} skipped", "added": added, "skipped": skipped}
